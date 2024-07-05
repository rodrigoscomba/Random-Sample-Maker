import os
import random
import shutil
from PIL import Image
import imagehash
import argparse
import openpyxl
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog, messagebox, Label, Entry, Button, Checkbutton, StringVar, BooleanVar, Radiobutton
import threading

def main_window(parent=None):
    # If parent is None, this is the main application window
    if parent is None:
        window = tk.Tk()
        window.title("Random Sample Maker")
    else:  # If parent is provided, attach to the parent window
        window = tk.Toplevel(parent)
        window.title("Random Sample Maker - Child Window")

    global start_button, progress_bar
    
    window.geometry("530x400")  # Adjusted window size
    print("START\n")
    main_frame = tk.Frame(window)
    main_frame.pack(padx=5, pady=5)
    main_frame.grid_columnconfigure(1, weight=1)

    # Arguments as global variables
    global keep_var, delete_var, ignore_folder_structure_var, size_var, duplicate_images_var,output_var, root_dir_var, dest_dir_var, num_samples_var, output_format_var

    # Initialize variables
    keep_var = BooleanVar(value=False)
    delete_var = BooleanVar(value=False)
    ignore_folder_structure_var = BooleanVar(value=False)
    size_var = StringVar()
    output_var = StringVar()
    root_dir_var = StringVar()
    dest_dir_var = StringVar()
    num_samples_var = StringVar()
    duplicate_images_var = BooleanVar(value=False)
    output_format_var = StringVar(value='PNG')

    # Function to update directory path label
    def update_label_path(label, path):
        label.config(text=path)

    # Create input fields with labels next to them
    widget_width = 35  # Width for widgets
    widget_vertical_pad = 5
    row = 0
    
    Label(main_frame, text="Delete copied images from root folder:", width=widget_width, anchor='w').grid(row=row, column=0, sticky='w', pady=(widget_vertical_pad, 0))
    Checkbutton(main_frame, text="", variable=delete_var).grid(row=row, column=1, sticky='w')
    row += 1
    
    Label(main_frame, text="Keep existing images in destination folder:", width=widget_width, anchor='w').grid(row=row, column=0, sticky='w', pady=(widget_vertical_pad, 0))
    Checkbutton(main_frame, text="", variable=keep_var).grid(row=row, column=1, sticky='w')
    row += 1
    
    Label(main_frame, text="Ignore directory structure in folder:", width=widget_width, anchor='w').grid(row=row, column=0, sticky='w', pady=(widget_vertical_pad, 0))
    Checkbutton(main_frame, text="", variable=ignore_folder_structure_var).grid(row=row, column=1, sticky='w')
    row += 1
    
    Label(main_frame, text="Duplicate images for unrepresented classes:", width=widget_width, anchor='w').grid(row=row, column=0, sticky='w', pady=(widget_vertical_pad, 0))
    Checkbutton(main_frame, text="", variable=duplicate_images_var).grid(row=row, column=1, sticky='w')
    row += 1
    
    Label(main_frame, text="Select output image type:", width=widget_width, anchor='w').grid(row=row, column=0, sticky='w', pady=(widget_vertical_pad, 0))
    formats_frame = tk.Frame(main_frame)
    formats_frame.grid(row=row, column=1, sticky='w')

    Radiobutton(formats_frame, text="PNG", variable=output_format_var, value='png').pack(side='left')
    Radiobutton(formats_frame, text="JPEG", variable=output_format_var, value='jpeg').pack(side='left')
    Radiobutton(formats_frame, text="BMP", variable=output_format_var, value='bmp').pack(side='left')
    Radiobutton(formats_frame, text="GIF", variable=output_format_var, value='gif').pack(side='left')
    Radiobutton(formats_frame, text="TIFF", variable=output_format_var, value='tiff').pack(side='left')
    row += 1

    Label(main_frame, text="Image Size (WIDTHxHEIGHT):", width=widget_width, anchor='w').grid(row=row, column=0, sticky='w', pady=(widget_vertical_pad, 0))
    size_entry = Entry(main_frame, textvariable=size_var, width=widget_width)
    size_entry.insert(0, "96x96")
    size_entry.grid(row=row, column=1, sticky='ew')
    row += 1

    Label(main_frame, text="Output Excel File Name:", width=widget_width, anchor='w').grid(row=row, column=0, sticky='w', pady=(widget_vertical_pad, 0))
    output_entry = Entry(main_frame, textvariable=output_var, width=widget_width)
    output_entry.insert(0, "output.xlsx")
    output_entry.grid(row=row, column=1, sticky='ew')
    row += 1

    def choose_directory(var, label):
        directory = filedialog.askdirectory()
        if directory:
            var.set(directory)
            update_label_path(label, directory)

    root_dir_label = Label(main_frame, text="", width=widget_width, anchor='w')
    dest_dir_label = Label(main_frame, text="", width=widget_width, anchor='w')

    Button(main_frame, text="Select Root Folder", command=lambda: choose_directory(root_dir_var, root_dir_label), width=widget_width).grid(row=row, column=0, columnspan=2, pady=(widget_vertical_pad, 5), sticky='ew')
    root_dir_label.grid(row=row+1, column=0, columnspan=2, sticky='ew')
    row += 2

    Button(main_frame, text="Select Destination Folder", command=lambda: choose_directory(dest_dir_var, dest_dir_label), width=widget_width).grid(row=row, column=0, columnspan=2, pady=(widget_vertical_pad, 5), sticky='ew')
    dest_dir_label.grid(row=row+1, column=0, columnspan=2, sticky='ew')
    row += 2

    Label(main_frame, text="Number of Samples:", width=widget_width, anchor='w').grid(row=row, column=0, sticky='w', pady=(widget_vertical_pad, 0))
    num_samples_entry = Entry(main_frame, textvariable=num_samples_var, width=widget_width)
    num_samples_entry.insert(0, "20")
    num_samples_entry.grid(row=row, column=1, sticky='ew')
    row += 1

    start_button = Button(main_frame, text="Make Random Sample", command=start_processing, width=widget_width)
    start_button.grid(row=row, column=0, columnspan=2, pady=(10, 0), sticky='ew')

    # Add progress bar
    progress_bar = ttk.Progressbar(main_frame, orient="horizontal", length=200, mode="determinate")
    progress_bar.grid(row=row+1, column=0, columnspan=2, pady=(10, 0), sticky='ew')
    
    window.mainloop()


def validate_args(args):
    """ Validate the input arguments. """

    if not os.path.isdir(args['rootdirectory']):
        messagebox.showerror("Error", "Invalid root directory.")
        return False
    if not os.path.isdir(args['destdirectory']):
        messagebox.showerror("Error", "Invalid destination directory.")
        return False
    if not args['size'].replace('x', '').isdigit() or 'x' not in args['size']:
        messagebox.showerror("Error", "Invalid image size format. Use WIDTHxHEIGHT.")
        return False
    if args['rootdirectory'] == args['destdirectory']:
        messagebox.showerror("Error", "Destination directory must not be the same as root directory.")
        return False
    try:
        num_samples = int(args['numbersamples'])
        if num_samples <= 0:
            raise ValueError
    except ValueError:
        messagebox.showerror("Error", "Number of samples must be a positive integer.")
        return False

    return True

def count_files_in_folder(folder):
    """ Counts the number of image files in the given folder """
    return len([file for file in os.listdir(folder) if file.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp', '.tiff'))])


def show_delete_warning():
    """ Show a warning dialog if delete option is selected """
    return messagebox.askokcancel("Warning", 
        "You have the option \"Delete copied images from root folder\" selected, this will delete all copied images from the original folder structure. This action cannot be undone.", 
        icon='warning', default='cancel', type='okcancel', detail="Press Ok to proceed.")

def start_processing():

    global start_button
    start_button.config(text="Running...", state="disabled")  # Update button text and disable it
    
    def threaded_processing():
        args = {
            'keep': keep_var.get(),
            'delete': delete_var.get(),
            'ignorefolderstructure': ignore_folder_structure_var.get(),
            'size': size_var.get(),
            'output': output_var.get(),
            'rootdirectory': root_dir_var.get(),
            'destdirectory': dest_dir_var.get(),
            'numbersamples': num_samples_var.get(),
            'makeduplicates': duplicate_images_var.get(),
            'outputformat': output_format_var.get()
        }

        # Check if delete option is selected and show warning
        if args['delete'] and not show_delete_warning():
            start_button.config(text="Make Random Sample", state="normal")
            return
        
        if validate_args(args):
            try:
                dorandomsamplemaker(args)
            except Exception as e:
                messagebox.showerror("Error", str(e))
            finally:
                start_button.config(text="Make Random Sample", state="normal") 
                progress_bar['value'] = 0
        else:
            start_button.config(text="Make Random Sample", state="normal")
    processing_thread = threading.Thread(target=threaded_processing)
    processing_thread.start()


def export_hierarchy_to_xlsx(folder_structure, workbook, parent_sheet, start_row, start_col, indent=0):
    """
    Recursively exports the folder structure to an Excel file with indentation.
    
    :param folder_structure: A nested dictionary representing the folder structure.
    :param workbook: An openpyxl workbook object to write to.
    :param parent_sheet: The active sheet in the workbook to write to.
    :param start_row: The row in the sheet to start writing to.
    :param start_col: The column in the sheet to start writing the folder name.
    :param indent: The current level of indentation (used for recursive calls).
    """
    if parent_sheet is None:
        parent_sheet = workbook.active
        parent_sheet.append(["Folder Path", "Image Count"])

    for folder, content in sorted(folder_structure.items()):
        if folder.startswith('_'):  # Skip the keys that are not folder names
            continue
        
        # Write the folder path and image count to the sheet with indentation
        parent_sheet.cell(row=start_row, column=start_col + indent, value=folder)
        if '_image_count' in folder_structure[folder]:
            parent_sheet.cell(row=start_row, column=start_col + indent + 1, value=folder_structure[folder]['_image_count'])
        
        # Recursively call this function for subdirectories
        if isinstance(content, dict):
            start_row = export_hierarchy_to_xlsx(content, workbook, parent_sheet, start_row + 1, start_col, indent + 1)
        else:
            start_row += 1
            
    return start_row

# Use this function to export your directory structure to an Excel file:
def export_folder_hierarchy(start_path, output_filename):
    folder_structure = build_folder_structure(start_path)
    workbook = openpyxl.Workbook()
    export_hierarchy_to_xlsx(folder_structure, workbook, None, 1, 1)
    output_path = os.path.join(start_path, output_filename)
    workbook.save(output_path)

def build_folder_structure(start_path):
    """
    Recursively builds a nested dictionary that represents the folder structure.
    """
    folder_structure = {}
    for item in os.listdir(start_path):
        item_path = os.path.join(start_path, item)
        if os.path.isdir(item_path):
            folder_structure[item] = build_folder_structure(item_path)
        else:
            if item.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp')):
                folder_structure.setdefault('_images', []).append(item)
    folder_structure['_image_count'] = len(folder_structure.get('_images', []))
    return folder_structure

def export_to_xlsx(image_counts, filename):
    """
    Export image counts to an Excel file.
    
    :param image_counts: A dictionary with keys as folder paths and values as image counts.
    :param filename: The filename for the Excel file.
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    sheet.append(["Folder Path", "Image Count"])
    for folder_path, count in image_counts.items():
        sheet.append([folder_path, count])

    workbook.save(filename)


def parse_args():
    parser = argparse.ArgumentParser(description='Copy and resize images.')
    parser.add_argument('-k', '--keep', action='store_true',
                        help='Keep existing images in the destination folder')
    parser.add_argument('-d', '--delete', action='store_true',
                    help='Delete original images after copying')
    parser.add_argument('-s', '--size', type=str, default="96x96",
                        help='Size to resize images, format WIDTHxHEIGHT (e.g., 128x128)')
    parser.add_argument('-o', '--output', type=str,
                        help='Specify the output Excel file name (e.g., output.xlsx)')
    parser.add_argument('-rd', '--rootdirectory', type=str, default=os.getcwd(),
                        help='Specify the root directory to build the folder structure (e.g., "C:\\myfolder")')
    parser.add_argument('-dd', '--destdirectory', type=str, default="../.",
                        help='Specify the destination directory to copy the images (e.g., "../MyDatasetSample")')
    parser.add_argument('-ns', '--numbersamples', type=str, default="20",
                        help='Specify the number of samples you wish to take from each folder (e.g., "20")')
    return parser.parse_args()

def clear_destination_folder(dest_folder, keep_images):
    if not keep_images and os.path.exists(dest_folder):
        try:
            for item in os.listdir(dest_folder):
                item_path = os.path.join(dest_folder, item)
                if os.path.isdir(item_path):
                    shutil.rmtree(item_path)
                else:
                    os.remove(item_path)
        except Exception as e:
            print(f"Failed to clear the destination folder: {dest_folder}. Reason: {e}")

def resize_and_crop_image(image_path, dest_path, output_size=(96, 96), output_format='PNG'):
    with Image.open(image_path) as img:
        
        img = img.convert("RGB")
        
        img_resized = img
        
        if img.size != output_size:
            # Calculate the crop size and offset to center the crop
            width, height = img.size
            crop_size = min(img.size)
            left = (width - crop_size) // 2
            top = (height - crop_size) // 2
            right = (width + crop_size) // 2
            bottom = (height + crop_size) // 2

            # Crop and resize the image
            img_cropped = img.crop((left, top, right, bottom))
            img_resized = img_cropped.resize(output_size, Image.Resampling.LANCZOS)

            # Convert to RGB if necessary
            if img_resized.mode in ("RGBA", "LA"):
                img_resized = img_resized.convert("RGB")

        img_resized.save(dest_path, format=output_format.upper())

def dorandomsamplemaker(args):
    output_size = tuple(map(int, args['size'].split('x')))
    root_directory = args['rootdirectory']
    dest_directory = args['destdirectory']
    number_samples = args['numbersamples']

    if os.path.abspath(root_directory) == os.path.abspath(dest_directory):
        return  # Exit the function

    # Only clear the destination folder if 'keep' is False.
    if not args['keep']:
        clear_destination_folder(dest_directory, args['keep'])
    
    file_count = 1
    
    total_files = sum([len(files) for _, _, files in os.walk(root_directory)])
    processed_files = 0

    for root, dirs, files in os.walk(root_directory):
        rel_path = os.path.relpath(root, root_directory)
        dest_path = dest_directory if args['ignorefolderstructure'] else os.path.join(dest_directory, rel_path)
        folder_hierarchy = rel_path.split(os.sep) if rel_path != "." else []
        os.makedirs(dest_path, exist_ok=True)

        if args['keep']:
            # If keeping files, start file_count based on what's already in the folder
            file_count = count_files_in_folder(dest_path) + 1

        # Call copy_random_images and update the file_count
        file_count = copy_random_images(root, dest_path, number_samples, args['keep'], output_size, file_count, args['delete'], args['makeduplicates'], args['outputformat'], folder_hierarchy)

        # If we're not ignoring the folder structure, reset file count to what's in the folder if 'keep' is True, or to 1 if 'keep' is False.
        if not args['ignorefolderstructure']:
            file_count = count_files_in_folder(dest_path) + 1 if args['keep'] else 1

        # Update processed files count and progress bar
        processed_files += len(files)
        progress = (processed_files / total_files) * 100
        progress_bar['value'] = progress
        
    if args['output']:
        export_folder_hierarchy(dest_directory, args['output'])

def generate_image_name(file_count, folder_hierarchy, output_format):
    """
    Generate a new image name based on the file count and folder hierarchy.
    
    :param file_count: The current count of files to generate a unique name.
    :param folder_hierarchy: A list of folder names from the root to the current folder.
    :return: A string representing the new image name.
    """
    hierarchy_string = "_".join(folder_hierarchy)
    return f"{file_count:04d}_{hierarchy_string}.{output_format}"

def are_images_duplicates(img_path1, img_path2, hash_func=imagehash.phash, hash_size=8):
    """
    Compare two images and return True if they are duplicates using perceptual hashing.
    
    :param img_path1: Path to the first image.
    :param img_path2: Path to the second image.
    :param hash_func: Hash function to use for image comparison (default: phash).
    :param hash_size: Hash size for image comparison (default: 8).
    :return: True if images are duplicates, otherwise False.
    """
    hash1 = hash_func(Image.open(img_path1), hash_size)
    hash2 = hash_func(Image.open(img_path2), hash_size)
    return hash1 == hash2

def remove_duplicates_from_list(source_folder, image_list, hash_func=imagehash.phash, hash_size=8):
    """
    Remove duplicate images from the given list of image paths using perceptual hashing.
    """
    unique_images = []
    hashes = set()

    for image_path in image_list:
        source_path = os.path.join(source_folder, image_path)
        try:
            with Image.open(source_path) as img:
                img = img.convert("RGB")
                image_hash = hash_func(img, hash_size)
                
                if image_hash not in hashes:
                    hashes.add(image_hash)
                    unique_images.append(image_path)
                else:
                    print(f"Duplicate found: {image_path}")
        except Exception as e:
            print(f"Error processing image {image_path}: {e}")
            continue

    return unique_images

def copy_random_images(source_folder, dest_folder, number_of_images, keep_images, output_size, file_count, delete_original, make_duplicates, output_image_type, folder_hierarchy):
    source_folder = r"{}".format(source_folder)
    dest_folder = r"{}".format(dest_folder)

    if not os.path.exists(source_folder):
        return file_count

    original_image_files = [file for file in os.listdir(source_folder) if file.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp'))]
    
    image_files = remove_duplicates_from_list(source_folder, original_image_files)
    
    selected_images = []
    
    # If 'number_of_images' is 'all' or more than available images, select all images.
    if len(image_files) > 0:
        if number_of_images.lower() == 'all':
            selected_images = image_files
        else:
            number_of_images = int(number_of_images)
            if len(image_files) < number_of_images:
                selected_images = image_files.copy()  # Copy the list of images
                # If 'makeduplicates' is set, duplicate images until we reach the required number
                if make_duplicates:
                    while len(selected_images) < number_of_images:
                        for image in image_files:
                            selected_images.append(image)
                            if len(selected_images) >= number_of_images:
                                break
            else:
                selected_images = random.sample(image_files, int(number_of_images))

    # Process and copy the selected images.
    for image in selected_images:
        source_path = os.path.join(source_folder, image)
        dest_filename = generate_image_name(file_count, folder_hierarchy, output_image_type)
        dest_path = os.path.join(dest_folder, dest_filename)

        if not keep_images or not os.path.exists(dest_path):
            resize_and_crop_image(source_path, dest_path, output_size, output_image_type)
            file_count += 1  # Increment file count after processing each image
        
        if delete_original:
            os.remove(source_path) 
            
    return file_count

def parse_command(command, default_size):
    try:
        parts = command.split(' ')
        if len(parts) < 2 or parts[0].lower() != "make":
            raise ValueError("Invalid command format.")

        number_of_images = parts[1]

        path_flag = "path="
        dest_flag = "dest="
        path_start = command.find(path_flag) + len(path_flag)
        dest_start = command.find(dest_flag) + len(dest_flag)
        source_folder = command[path_start:].split('"')[1]
        dest_folder = command[dest_start:].split('"')[1]
        
        size_flag = "size="
        if size_flag in command:
            size_start = command.find(size_flag) + len(size_flag)
            size_str = command[size_start:].split(' ')[0].replace('"', '')
            size_parts = size_str.split('x')
            size = (int(size_parts[0]), int(size_parts[1]))
        else:
            size = default_size

        return source_folder, dest_folder, number_of_images, size
    except Exception as e:
        #print(f"Error parsing command: {e}")
        return None, None, None, default_size
   
if __name__ == '__main__':
    main_window()